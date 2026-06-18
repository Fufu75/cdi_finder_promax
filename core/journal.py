"""Ajout d'une ligne dans le tableau de suivi Excel."""
from datetime import date
from pathlib import Path
import openpyxl

CHEMIN_DEFAUT = "suivi.xlsx"

COLONNES = [
    "Date",
    "Poste",
    "Entreprise",
    "Lieu",
    "Type contrat",
    "Statut",
    "CV généré",
    "Lettre générée",
    "Lien offre",
    "Notes",
]


def _initialiser(chemin: str) -> openpyxl.Workbook:
    """Crée le fichier Excel avec l'en-tête si inexistant."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi candidatures"

    # En-tête
    ws.append(COLONNES)

    # Mise en forme de l'en-tête
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # Largeurs de colonnes
    largeurs = [12, 30, 25, 20, 14, 16, 14, 16, 40, 30]
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = largeur

    Path(chemin).parent.mkdir(parents=True, exist_ok=True)
    wb.save(chemin)
    return wb


def enregistrer_candidature(
    offre_analysee: dict,
    chemin_cv: str = "",
    chemin_lettre: str = "",
    lien_offre: str = "",
    notes: str = "",
    chemin_excel: str = CHEMIN_DEFAUT,
):
    """Ajoute une ligne dans suivi.xlsx. Crée le fichier si nécessaire."""
    if not Path(chemin_excel).exists():
        _initialiser(chemin_excel)

    wb = openpyxl.load_workbook(chemin_excel)
    ws = wb.active

    ws.append([
        date.today().isoformat(),
        offre_analysee.get("poste", ""),
        offre_analysee.get("entreprise", ""),
        offre_analysee.get("lieu", ""),
        offre_analysee.get("type_contrat", ""),
        "À envoyer",
        "Oui" if chemin_cv else "Non",
        "Oui" if chemin_lettre else "Non",
        lien_offre,
        notes,
    ])

    wb.save(chemin_excel)
